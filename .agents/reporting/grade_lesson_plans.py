import os
import json
import random
from pathlib import Path
import subprocess
import sys
from datetime import datetime

WORKSPACE_ROOT = Path(__file__).resolve().parents[2]
OUTPUT_FOLDER = WORKSPACE_ROOT / "Output giáo án docs"
REPORTING_DIR = WORKSPACE_ROOT / ".agents" / "reporting"
MD_FOLDER = OUTPUT_FOLDER / "markdown"
REPORT_PATH = REPORTING_DIR / "grading_report.json"

def evaluate_lesson_plan(md_content: str, filename: str) -> dict:
    """
    Evaluates a lesson plan against the rubric.
    In a real scenario, this would call an LLM (e.g. Gemini).
    Here we provide a robust mock evaluation with some random variation
    to demonstrate the alerting capabilities.
    """
    # Simple heuristic to determine "bad" lesson plans for testing
    # If filename has "Math", we'll sometimes make it fail.
    
    score_data = {
        "Data Extraction (15)": 15,
        "Montessori Standards (20)": 20,
        "Age Appropriateness (15)": 15,
        "Scaffolding (15)": 15,
        "Safety (10)": 10,
        "Engagement (15)": 15,
        "Formatting (10)": 10
    }
    
    failed_items = []
    suggestions = []
    
    # Let's mock a failure for demonstration if 'Math' or 'Lesson 2' is in the filename
    if "Math" in filename or random.random() < 0.3:
        score_data["Age Appropriateness (15)"] = 0
        score_data["Engagement (15)"] = 0
        score_data["Safety (10)"] = 0
        score_data["Scaffolding (15)"] = 0
        
        failed_items.append("Độ An Toàn Lớp Học (Cảnh báo đỏ)")
        failed_items.append("Chuẩn Phương Pháp Dạy Theo Độ Tuổi")
        failed_items.append("Tính Tò Mò, Hứng Thú")
        
        suggestions.append("Cảnh báo: Thiếu ghi chú an toàn (Safety Note) cho hoạt động sử dụng kéo hoặc vật nhỏ.")
        suggestions.append("Hoạt động quá tĩnh, cần bổ sung TPR (Total Physical Response) và các bài hát khởi động.")
        suggestions.append("Lesson plan thiếu Trò chơi/vận động/thiếu tính sáng tạo, học sinh phải ngồi ghế quá lâu.")
        
    elif "Show" in filename:
        # A mediocre one
        score_data["Montessori Standards (20)"] = 10
        score_data["Data Extraction (15)"] = 10
        failed_items.append("Tiêu chuẩn Montessori (Thiếu Control of Error)")
        suggestions.append("Chưa có cơ chế để học sinh tự kiểm tra lỗi (Self-correction). Cần bổ sung Word Mat úp sấp.")
    
    total_score = sum(score_data.values())
    
    return {
        "filename": filename,
        "total_score": total_score,
        "scores": score_data,
        "failed_items": failed_items,
        "suggestions": suggestions,
        "status": "PASS" if total_score >= 50 else "FAIL"
    }

def main():
    print(f"[GRADER] Looking for lesson plans in {MD_FOLDER}")
    
    if not MD_FOLDER.exists():
        print(f"[GRADER] Output directory not found. Please run the pipeline first.")
        return
        
    md_files = list(MD_FOLDER.glob("*.md"))
    if not md_files:
        print("[GRADER] No markdown files found to grade.")
        return
        
    print(f"[GRADER] Found {len(md_files)} lesson plans to grade.")
    
    report_data = []
    has_failures = False
    
    for md_file in md_files:
        content = md_file.read_text(encoding="utf-8")
        result = evaluate_lesson_plan(content, md_file.name)
        report_data.append(result)
        
        if result["status"] == "FAIL":
            has_failures = True
            print(f"  ❌ FAIL: {md_file.name} (Score: {result['total_score']})")
        else:
            print(f"  ✅ PASS: {md_file.name} (Score: {result['total_score']})")
            
    # Save the report as JSON
    REPORTING_DIR.mkdir(parents=True, exist_ok=True)
    with open(REPORT_PATH, 'w', encoding='utf-8') as f:
        json.dump(report_data, f, ensure_ascii=False, indent=2)
        
    print(f"\n[GRADER] Grading complete. Report saved to {REPORT_PATH.name}")

    # Read schedule metadata if available to fill columns accurately
    sched_path = OUTPUT_FOLDER / "_schedule.json"
    age_group = "3-4"
    sched_map = {}
    if sched_path.exists():
        try:
            with open(sched_path, 'r', encoding='utf-8') as sf:
                s_data = json.load(sf)
                age_group = s_data.get("age", "3-4")
                for slot in s_data.get("schedule", []):
                    src = slot.get("source")
                    if src:
                        sched_map[src] = slot
        except Exception as e:
            print(f"[GRADER] Warning reading schedule metadata: {e}")

    # Get current timestamp
    current_ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Save the report as CSV in Output giáo án docs using utf-8-sig to prevent Excel corruption
    import csv
    csv_path = OUTPUT_FOLDER / "grading_report.csv"
    
    headers = [
        "Timestamp", "Week / Topic", "File Name", "Age Group", "Status", "Total Score",
        "Data Extraction (15)", "Montessori Standards (20)", "Age Appropriateness (15)",
        "Scaffolding (15)", "Safety (10)", "Engagement (15)", "Formatting (10)",
        "Failed Items", "AI Suggestions", "Docx Link", "PDF Link", "Reviewer", "Manager Notes"
    ]

    rows_data = []
    for r in report_data:
        s = r["scores"]
        stem = Path(r["filename"]).stem
        slot_info = sched_map.get(stem) or {}
        week_topic = f"{slot_info.get('week', 'Week 1')} - {slot_info.get('topic', 'General Topic')}"
        
        rows_data.append([
            current_ts,
            week_topic,
            r["filename"],
            age_group,
            r["status"],
            r["total_score"],
            s.get("Data Extraction (15)", 0),
            s.get("Montessori Standards (20)", 0),
            s.get("Age Appropriateness (15)", 0),
            s.get("Scaffolding (15)", 0),
            s.get("Safety (10)", 0),
            s.get("Engagement (15)", 0),
            s.get("Formatting (10)", 0),
            "; ".join(r["failed_items"]),
            "; ".join(r["suggestions"]),
            "", # Docx Link
            "", # PDF Link
            "", # Reviewer
            ""  # Manager Notes
        ])

    with open(csv_path, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows_data)
        
    print(f"[GRADER] CSV Report saved to {csv_path} (UTF-8 with BOM)")

    # Write formatted XLSX using openpyxl
    try:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            print("[GRADER] Installing openpyxl for premium Excel formatting...")
            os.system(f"{sys.executable} -m pip install openpyxl -q")
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

        xlsx_path = OUTPUT_FOLDER / "grading_report.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Grading Report"
        
        # Enable gridlines
        ws.views.sheetView[0].showGridLines = True
        
        # 1. Super Header
        ws.merge_cells("A1:S1")
        ws["A1"] = "HỆ THỐNG ĐÁNH GIÁ CHẤT LƯỢNG GIÁO ÁN TỰ ĐỘNG BẰNG AI"
        ws["A1"].font = Font(name="Inter", size=16, bold=True, color="FFFFFF")
        ws["A1"].fill = PatternFill(start_color="3050A2", end_color="3050A2", fill_type="solid") # Cyan Cobalt Blue
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40
        
        # 2. Table Headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = Font(name="Inter", size=11, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="29B198", end_color="29B198", fill_type="solid") # Mountain Meadow
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[2].height = 28
        
        # Freeze rows 1 and 2
        ws.freeze_panes = "A3"
        
        # Style Definitions
        border_side = Side(border_style="thin", color="D3D3D3")
        thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
        
        fill_pass = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Light Green
        font_pass = Font(name="Inter", size=10, bold=True, color="375623")
        
        fill_fail = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # Light Red
        font_fail = Font(name="Inter", size=10, bold=True, color="C00000")
        
        # 3. Data Rows
        for r_idx, row_val in enumerate(rows_data, 3):
            ws.row_dimensions[r_idx].height = 24
            for c_idx, val in enumerate(row_val, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.font = Font(name="Inter", size=10)
                cell.border = thin_border
                
                # Alignments
                if c_idx in [1, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13]:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                
                # Status formatting
                if c_idx == 5:
                    if val == "PASS":
                        cell.fill = fill_pass
                        cell.font = font_pass
                    else:
                        cell.fill = fill_fail
                        cell.font = font_fail
                        
                # Wrap Text for Suggesions / Failed Items
                if c_idx in [14, 15, 19]:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            # Skip merged title row when calculating width
            for cell in col[1:]:
                val_str = str(cell.value or '')
                if cell.alignment.wrap_text:
                    max_len = max(max_len, min(len(val_str), 30))
                else:
                    max_len = max(max_len, len(val_str))
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
        wb.save(xlsx_path)
        print(f"[GRADER] Styled XLSX Report saved to {xlsx_path}")
    except Exception as e:
        print(f"[GRADER] Warning generating styled Excel: {e}")

    # Trigger email alert if there are failures
    if has_failures:
        print("[GRADER] Failures detected. Triggering Email Alert System...")
        alert_script = REPORTING_DIR / "send_email_alert.py"
        subprocess.run([sys.executable, str(alert_script)])

if __name__ == "__main__":
    main()
